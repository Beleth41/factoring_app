import streamlit as st
import pandas as pd
from openpyxl import load_workbook



def read_excel_with_progress(file, sheet_name, header_row=1):
    # header_row correspond exactement au numéro de ligne Excel (1-based)
    wb = load_workbook(filename=file, read_only=True, data_only=True)
    ws = wb[sheet_name]

    max_row = ws.max_row

    # Récupérer l'entête à la ligne spécifiée (header_row)
    header = [cell.value for cell in ws[header_row]]

    data = []
    # Calculer le nombre total de lignes à lire (après l'entête)
    total_rows_to_read = max_row - header_row
    progress_bar = st.progress(0)
    status_text = st.empty()

    # Lire les données à partir de la ligne suivant l'entête
    for i, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=max_row, values_only=True), start=1):
        data.append(row)

        if i % 10 == 0 or i == total_rows_to_read:
            progress = i / total_rows_to_read
            progress_bar.progress(progress)
            status_text.text(f"Lecture de la ligne {i} / {total_rows_to_read}")

    # Construire le DataFrame
    df = pd.DataFrame(data, columns=header)

    # Nettoyer les colonnes
    df.columns = [str(col).strip() if col is not None else f"Unnamed_{i}" 
                  for i, col in enumerate(df.columns)]

    return df


# Streamlit app
st.title("Page d'accueil")

if "df" not in st.session_state:
    st.session_state.df = None

upload_file = st.file_uploader("Télécharger un fichier Excel", type=["xlsx"])

if upload_file and st.session_state.df is None:
    wb = load_workbook(upload_file, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    selected_sheet = st.selectbox("Sélectionnez une feuille", sheet_names)
    header_row = st.number_input("Numéro de la ligne d'entête (5 = ligne 5 d'Excel)", min_value=1, value=5, step=1)

    if st.button("Charger cette feuille"):
        try:
            df = read_excel_with_progress(upload_file, selected_sheet, header_row=int(header_row))
            st.session_state.df = df
            st.success(f"Fichier chargé avec succès ({selected_sheet}) !")
        except Exception as e:
            st.error(f"Erreur lors du chargement du fichier: {e}")

if st.session_state.df is not None:
    st.markdown("---")
    st.subheader("Choisissez votre type d'analyse")

    # Détecter automatiquement le type de base de données
    df = st.session_state.df
    
    # Colonnes de la première base (originale)
    base1_columns = ['TIRES', 'Debtor Number', 'EntryAmountSAC', 'RUB']
    
    # Colonnes de la deuxième base (alternative)
    base2_columns = ['Entry Amount SAC', 'Rubrique', 'MVT', 'ledger item id']
    
    # Vérifier quelle base correspond
    has_base1 = all(col in df.columns for col in base1_columns)
    has_base2 = all(col in df.columns for col in base2_columns)
    
    if has_base1 and not has_base2:
        st.info("**Base de données détectée :** Format original avec colonnes TIRES, RUB, etc.")
        # Stocker le type de base dans session_state
        st.session_state.base_type = "original"
        
        option = st.radio(
            "Sélectionnez une analyse à effectuer :",
            ["Analyse par adhérent", "Analyse par Tirés"]
        )
        
        if st.button("Ouvrir l'analyse"):
            if option == "Analyse par adhérent":
                st.switch_page("pages/1_Analyse_Adhérent.py")
            elif option == "Analyse par Tirés":
                st.switch_page("pages/2_Analyse_TIRES.py")
                
    elif has_base2 and not has_base1:
        st.info("**Base de données détectée :** Format alternatif avec colonnes Rubrique, MVT, etc.")
        # Stocker le type de base dans session_state
        st.session_state.base_type = "alternative"
        
        option = st.radio(
            "Sélectionnez une analyse à effectuer :",
            ["Analyse générale", "Analyse par adhérent (base alt)", "Recherche par client"]
        )
        
        if st.button("Ouvrir l'analyse"):
            if option == "Analyse générale":
                st.switch_page("pages/3_Analyse_Generale.py")
            elif option == "Analyse par adhérent (base alt)":
                st.switch_page("pages/3_Analyse_Generale.py")  # Même page avec fonctionnalité de recherche
            elif option == "Recherche par client":
                st.switch_page("pages/3_Analyse_Generale.py")  # Même page, section recherche
                
    elif has_base1 and has_base2:
        st.warning("**Colonnes mixtes détectées.** Veuillez choisir le format de votre base :")
        format_choice = st.radio(
            "Choisir le format d'analyse :",
            ["Format original (TIRES, RUB)", "Format alternatif (Rubrique, MVT)"]
        )
        
        if format_choice == "Format original (TIRES, RUB)":
            # Stocker le type de base dans session_state
            st.session_state.base_type = "original"
            
            option = st.radio(
                "Sélectionnez une analyse à effectuer :",
                ["Analyse par adhérent", "Analyse par Tirés"]
            )
            
            if st.button("Ouvrir l'analyse"):
                if option == "Analyse par adhérent":
                    st.switch_page("pages/1_Analyse_Adhérent.py")
                elif option == "Analyse par Tirés":
                    st.switch_page("pages/2_Analyse_TIRES.py")
        else:
            # Stocker le type de base dans session_state
            st.session_state.base_type = "alternative"
            
            option = st.radio(
                "Sélectionnez une analyse à effectuer :",
                ["Analyse générale"]
            )
            
            if st.button("Ouvrir l'analyse"):
                st.switch_page("pages/3_Analyse_Generale.py")
                
    else:
        st.error("**Format de base de données non reconnu.**")
        st.write("**Colonnes détectées dans votre fichier :**")
        st.write(", ".join(df.columns.tolist()))
        st.write("")
        st.write("**Formats supportés :**")
        st.write("- **Format 1 :** TIRES, Debtor Number, EntryAmountSAC, RUB, etc.")
        st.write("- **Format 2 :** Entry Amount SAC, Rubrique, MVT, ledger item id, etc.")

else:
    st.info("Veuillez charger un fichier pour continuer.")